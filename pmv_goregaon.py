from pythermalcomfort.models import pmv
from pythermalcomfort.models import pmv_ppd
from pythermalcomfort.utilities import v_relative, clo_dynamic
import pandas as pd
import xlwt 
#from xlwt import Workbook
from openpyxl import load_workbook
# Workbook is created 
#wb = Workbook()
#str1=""
#d1=list('Sheet')
#sheet1 = wb.add_sheet(str1.join(d1),cell_overwrite_ok=True)
filepath="IDBI_MUM_goregoan_main_lobby_AC1_5th_Jun_2023_to_14th_June_2023.xlsx"
wb = load_workbook(filepath)
sheet = wb.active
sheet.insert_cols(idx=18)
dataframe1 = pd.read_excel(filepath)
#print(dataframe1)
#Duct Length:L=72 cm,W=7 cm
#Indoor AC Fan Motor:1100 rpm,radius=0.0047 m(1 ton)
#Indoor AC Fan Motor:1450 rpm,radius=0.0047 m(1 ton) v=0.71
#v(m/s)=2*pi*r*N(rpm)/60 Ref:https://lucidar.me/en/unit-converter/revolutions-per-minute-to-meters-per-second/
c1 = sheet.cell(row = 1, column = 18)
c1.value="PMV"
#print(len(dataframe1['TEMP']))
#print(v_relative(v=0.03048,met=1))
for i in range(0,len(dataframe1['pl.Tempc']),1):
  tdb = float(dataframe1['pl.Tempc'][i])
  print(tdb)
  tr = tdb
  rh = float(dataframe1['pl.Humidity'][i])
  print(rh)
  m=dataframe1['m'][i]
  print(m)
  if m=="C8:F0:9E:29:43:91" or m=="C8:F0:9E:29:3F:99" or m=="C8:F0:9E:29:3C:15" or m=="C8:F0:9E:29:41:4D" or m=="C8:F0:9E:28:00:CD":
    v = 0.002442307692
  elif m=="C8:F0:9E:29:3F:E1":
    v=0.0127
  elif m=="C8:F0:9E:29:42:01":
    v=0.03762962963
  elif m=="C8:F0:9E:29:40:95" or m=="C8:F0:9E:28:21:D5":
    v=0.01905
  elif m=="C8:F0:9E:29:3E:59" or m=="C8:F0:9E:29:3D:F1":
    v=0.02116666667
  elif m=="C8:F0:9E:29:40:19" or m=="C8:F0:9E:29:41:F5":
    v=0.04064
  print(v)
  met = 1#Seated,quiet:1.0 met
  clo = 0.61#Trousers,long-sleeve shirt
  # calculate relative air speed
  v_r = v_relative(v=v, met=met)
  #print(v_r)
  # calculate dynamic clothing
  clo_d = clo_dynamic(clo=clo, met=met)
  #results = pmv(tdb=tdb, tr=tr, vr=v_r, rh=rh, met=met, clo=clo_d)
  results=pmv_ppd(tdb=tdb, tr=tr, vr=v_r, rh=rh, met=met, clo=clo, standard="ASHRAE")
  #print(results)
  c1=sheet.cell(row=i+2,column=18)
  c1.value=results['pmv']
  print(results['pmv'])
wb.save(filepath)

