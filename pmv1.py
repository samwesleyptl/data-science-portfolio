from pythermalcomfort.models import pmv
from pythermalcomfort.models import pmv_ppd
from pythermalcomfort.utilities import v_relative, clo_dynamic
import pandas as pd
import xlwt 
#from xlwt import Workbook
from openpyxl import load_workbook
# Workbook is created 
#wb = Workbook()
#str1=""
#d1=list('Sheet')
#sheet1 = wb.add_sheet(str1.join(d1),cell_overwrite_ok=True)
filepath="Kotak_AHM_Narol_Cross_AC2_7th_June_2023_to_26th_June_2023.xlsx"
wb = load_workbook(filepath)
sheet = wb.active
sheet.insert_cols(idx=18)
dataframe1 = pd.read_excel(filepath)
#print(dataframe1)
#Duct Length:L=72 cm,W=7 cm
#Indoor AC Fan Motor:1100 rpm,radius=0.0047 m(1 ton)
#Indoor AC Fan Motor:1450 rpm,radius=0.0047 m(1 ton) v=0.71
#v(m/s)=2*pi*r*N(rpm)/60 Ref:https://lucidar.me/en/unit-converter/revolutions-per-minute-to-meters-per-second/
c1 = sheet.cell(row = 1, column = 18)
c1.value="PMV"
#print(len(dataframe1['TEMP']))
#print(v_relative(v=0.03048,met=1))

for i in range(0,len(dataframe1['TEMP']),1):
  tdb = float(dataframe1['TEMP'][i])
  print(tdb)
  tr = tdb
  rh = float(dataframe1['HUMIDITY'][i])
  print(rh)
  m=dataframe1['MAC ID'][i]
  print(m)
  if m=="C8:F0:9E:29:42:D1" or m=="C8:F0:9E:29:42:ED" or m=="C8:F0:9E:29:41:85" or m=="C8:F0:9E:29:3F:3D" or m=="C8:F0:9E:29:3C:B9" or m=="C8:F0:9E:29:3D:C1" or m=="C8:F0:9E:29:41:11":
    v = 0.003256410256
  elif m=="C8:F0:9E:28:13:21" or m=="C8:F0:9E:29:43:3D":
    v=0.01016
  elif m=="C8:F0:9E:29:3D:81":
    v=0.0254
  elif m=="C8:F0:9E:29:43:09" or m=="C8:F0:9E:27:FB:39":
    v=0.0635
  elif m=="C8:F0:9E:29:42:F1":
    v=0.04064
  elif m=="C8:F0:9E:29:43:19" or m=="C8:F0:9E:28:01:4D":
    v=0.03048
  elif m=="C8:F0:9E:29:42:A9" or m=="C8:F0:9E:27:E8:C5":
    v=0.0254
  print(v)
  met = 1#Seated,quiet:1.0 met
  clo = 0.61#Trousers,long-sleeve shirt
  # calculate relative air speed
  v_r = v_relative(v=v, met=met)
  #print(v_r)
  # calculate dynamic clothing
  clo_d = clo_dynamic(clo=clo, met=met)
  #results = pmv(tdb=tdb, tr=tr, vr=v_r, rh=rh, met=met, clo=clo_d)
  results=pmv_ppd(tdb=tdb, tr=tr, vr=v_r, rh=rh, met=met, clo=clo, standard="ASHRAE")
  #print(results)
  c1=sheet.cell(row=i+2,column=18)
  c1.value=results['pmv']
  print(results['pmv'])
'''
for i in range(0,len(dataframe1['pl.Tempc']),1):
  tdb = float(dataframe1['pl.Tempc'][i])
  print(tdb)
  tr = tdb
  rh = float(dataframe1['pl.Humidity'][i])
  print(rh)
  m=dataframe1['m'][i]
  if m=="C8:F0:9E:29:42:D1" or m=="C8:F0:9E:29:42:ED" or m=="C8:F0:9E:29:41:85" or m=="C8:F0:9E:29:3F:3D" or m=="C8:F0:9E:29:3C:B9" or m=="C8:F0:9E:29:3D:C1" or m=="C8:F0:9E:29:41:11":
    v = 0.003256410256
  elif m=="C8:F0:9E:28:13:21" or m=="C8:F0:9E:29:43:3D":
    v=0.01016
  elif m=="C8:F0:9E:29:3D:81":
    v=0.0254
  elif m=="C8:F0:9E:29:43:09" or m=="C8:F0:9E:27:FB:39":
    v=0.0635
  elif m=="C8:F0:9E:29:42:F1":
    v=0.04064
  elif m=="C8:F0:9E:29:43:19" or m=="C8:F0:9E:28:01:4D":
    v=0.03048
  print(v)
  met = 1#Seated,quiet:1.0 met
  clo = 0.61#Trousers,long-sleeve shirt
  # calculate relative air speed
  v_r = v_relative(v=v, met=met)
  #print(v_r)
  # calculate dynamic clothing
  clo_d = clo_dynamic(clo=clo, met=met)
  #results = pmv(tdb=tdb, tr=tr, vr=v_r, rh=rh, met=met, clo=clo_d)
  results=pmv_ppd(tdb=tdb, tr=tr, vr=v_r, rh=rh, met=met, clo=clo, standard="ASHRAE")
  #print(results)
  c1=sheet.cell(row=i+2,column=18)
  c1.value=results['pmv']
  print(results['pmv'])
'''
wb.save(filepath)

